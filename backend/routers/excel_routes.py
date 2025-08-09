#!/usr/bin/env python3
"""
Excel Import/Export routes for Medical Management System
Supports exporting and importing data for all major sections
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
import os
import jwt
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import uuid
import io
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]

# JWT Configuration  
JWT_SECRET_KEY = "your-secret-key-change-in-production"
JWT_ALGORITHM = "HS256"

# Security
security = HTTPBearer()

# Create router
router = APIRouter(prefix="/api/excel", tags=["excel-operations"])

def verify_jwt_token(token: str) -> dict:
    """Verify and decode JWT token"""
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get current user from JWT token"""
    try:
        token = credentials.credentials
        payload = verify_jwt_token(token)
        return payload
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Authentication failed: {str(e)}")

def create_excel_template(data_type: str) -> Workbook:
    """Create Excel template with sample data"""
    wb = Workbook()
    ws = wb.active
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    templates = {
        "clinics": {
            "sheet_name": "العيادات - Clinics",
            "headers": [
                "اسم العيادة", "هاتف العيادة", "بريد العيادة", "اسم الطبيب", "هاتف الطبيب",
                "عنوان العيادة", "رمز الخط", "رمز المنطقة", "تصنيف العيادة", "التصنيف الائتماني", "ملاحظات"
            ],
            "sample_data": [
                ["عيادة د. أحمد محمد", "01234567890", "clinic@example.com", "د. أحمد محمد", "01111111111",
                 "شارع النيل، المعادي، القاهرة", "CN", "NC", "class_a", "green", "عيادة متميزة"],
                ["عيادة د. سارة أحمد", "01987654321", "sara@clinic.com", "د. سارة أحمد", "01222222222",
                 "شارع الجمهورية، وسط البلد", "ALX", "MB", "class_b", "yellow", "عيادة جيدة"]
            ]
        },
        "users": {
            "sheet_name": "المستخدمين - Users",
            "headers": [
                "اسم المستخدم", "الاسم الكامل", "كلمة المرور", "الدور", "البريد الإلكتروني", 
                "رمز الخط", "رمز المنطقة", "رمز المدير", "نشط"
            ],
            "sample_data": [
                ["ahmed.mohamed", "أحمد محمد علي", "password123", "medical_rep", "ahmed@company.com",
                 "CN", "NC", "admin-001", "نعم"],
                ["sara.ahmed", "سارة أحمد محمود", "password456", "accounting", "sara@company.com",
                 "ALX", "MB", "admin-001", "نعم"]
            ]
        },
        "orders": {
            "sheet_name": "الطلبات - Orders", 
            "headers": [
                "رقم الطلب", "رمز العيادة", "رمز المندوب", "رمز المنتج", "الكمية", "السعر",
                "إجمالي المبلغ", "حالة الطلب", "تاريخ الطلب", "ملاحظات"
            ],
            "sample_data": [
                ["ORD-001", "clinic-001", "rep-001", "PROD-001", "10", "50.00",
                 "500.00", "confirmed", "2025-01-15", "طلب عادي"],
                ["ORD-002", "clinic-002", "rep-002", "PROD-002", "5", "75.00",
                 "375.00", "pending", "2025-01-16", "طلب مستعجل"]
            ]
        },
        "debts": {
            "sheet_name": "المديونية - Debts",
            "headers": [
                "رقم المديونية", "رمز العيادة", "رمز المندوب", "المبلغ الأصلي", "المبلغ المتبقي",
                "تاريخ الاستحقاق", "حالة المديونية", "نوع المديونية", "ملاحظات"
            ],
            "sample_data": [
                ["DEBT-001", "clinic-001", "rep-001", "1000.00", "500.00",
                 "2025-02-15", "partially_paid", "invoice", "دين من فاتورة"],
                ["DEBT-002", "clinic-002", "rep-002", "750.00", "750.00",
                 "2025-03-01", "outstanding", "credit", "دين ائتماني"]
            ]
        },
        "payments": {
            "sheet_name": "التحصيل - Payments",
            "headers": [
                "رقم المدفوعة", "رقم المديونية", "رمز العيادة", "رمز المندوب", "المبلغ المدفوع",
                "طريقة الدفع", "تاريخ الدفع", "رقم الإيصال", "ملاحظات"
            ],
            "sample_data": [
                ["PAY-001", "DEBT-001", "clinic-001", "rep-001", "500.00",
                 "cash", "2025-01-20", "REC-001", "دفعة نقدية"],
                ["PAY-002", "DEBT-002", "clinic-002", "rep-002", "250.00",
                 "bank_transfer", "2025-01-25", "REC-002", "حوالة بنكية"]
            ]
        }
    }
    
    template = templates.get(data_type)
    if not template:
        raise HTTPException(status_code=400, detail=f"Unsupported data type: {data_type}")
    
    ws.title = template["sheet_name"]
    
    # Add headers
    for col, header in enumerate(template["headers"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Auto-adjust column width
        ws.column_dimensions[get_column_letter(col)].width = len(header) + 5
    
    # Add sample data
    for row, sample_row in enumerate(template["sample_data"], 2):
        for col, value in enumerate(sample_row, 1):
            ws.cell(row=row, column=col, value=value)
    
    return wb

@router.get("/template/{data_type}")
async def download_template(data_type: str, current_user: dict = Depends(get_current_user)):
    """Download Excel template with sample data for import"""
    try:
        # Check permissions
        if current_user.get("role") not in ["admin", "gm", "manager", "accounting"]:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        
        # Create template
        wb = create_excel_template(data_type)
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # URL-safe filenames without Arabic characters
        filename_mapping = {
            "clinics": "clinics_template",
            "users": "users_template", 
            "orders": "orders_template",
            "debts": "debts_template",
            "payments": "payments_template"
        }
        
        filename = f"{filename_mapping.get(data_type, data_type)}_template.xlsx"
        
        # Get the content as bytes
        content = output.getvalue()
        
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating template: {str(e)}")

@router.get("/export/{data_type}")
async def export_data(data_type: str, current_user: dict = Depends(get_current_user)):
    """Export data to Excel"""
    try:
        # Check permissions
        if current_user.get("role") not in ["admin", "gm", "manager", "accounting"]:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        
        # Collection mapping
        collections = {
            "clinics": db.clinics,
            "users": db.users,
            "orders": db.orders,
            "debts": db.debts,
            "payments": db.payments
        }
        
        if data_type not in collections:
            raise HTTPException(status_code=400, detail=f"Unsupported data type: {data_type}")
        
        # Get data from database
        data = []
        async for item in collections[data_type].find({}, {"_id": 0}):
            data.append(item)
        
        if not data:
            raise HTTPException(status_code=404, detail=f"No {data_type} data found")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        
        # Arabic sheet names
        sheet_names = {
            "clinics": "بيانات العيادات",
            "users": "بيانات المستخدمين", 
            "orders": "بيانات الطلبات",
            "debts": "بيانات المديونية",
            "payments": "بيانات التحصيل"
        }
        
        ws.title = sheet_names.get(data_type, data_type)
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        if data:
            # Add headers from first item keys
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Add data
            for row, item in enumerate(data, 2):
                for col, key in enumerate(headers, 1):
                    value = item.get(key, "")
                    # Convert datetime objects to strings
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, list):
                        value = ", ".join(str(v) for v in value)
                    elif isinstance(value, dict):
                        value = json.dumps(value, ensure_ascii=False)
                    ws.cell(row=row, column=col, value=str(value))
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # URL-safe filenames without Arabic characters
        filename_mapping = {
            "clinics": "clinics_export",
            "users": "users_export",
            "orders": "orders_export", 
            "debts": "debts_export",
            "payments": "payments_export"
        }
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_mapping.get(data_type, data_type)}_{timestamp}.xlsx"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_mapping.get(data_type, data_type)}_{timestamp}.xlsx"
        
        # Get the content as bytes
        content = output.getvalue()
        
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting data: {str(e)}")

@router.post("/import/{data_type}")
async def import_data(
    data_type: str,
    file: UploadFile = File(...),
    import_mode: str = Form("append"),  # "append" or "overwrite"
    current_user: dict = Depends(get_current_user)
):
    """Import data from Excel file"""
    try:
        # Check permissions
        if current_user.get("role") not in ["admin", "gm"]:
            raise HTTPException(status_code=403, detail="Only admin and GM can import data")
        
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
        
        # Collection mapping
        collections = {
            "clinics": db.clinics,
            "users": db.users,
            "orders": db.orders,
            "debts": db.debts,
            "payments": db.payments
        }
        
        if data_type not in collections:
            raise HTTPException(status_code=400, detail=f"Unsupported data type: {data_type}")
        
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers:
            raise HTTPException(status_code=400, detail="No headers found in Excel file")
        
        # Read data rows
        imported_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):  # Skip empty rows
                row_data = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) and row[i] is not None else ""
                    row_data[header] = str(value).strip() if value else ""
                
                # Add metadata
                row_data["id"] = str(uuid.uuid4())
                row_data["imported_at"] = datetime.utcnow().isoformat()
                row_data["imported_by"] = current_user.get("user_id")
                
                imported_data.append(row_data)
        
        if not imported_data:
            raise HTTPException(status_code=400, detail="No data rows found in Excel file")
        
        # Handle import mode
        collection = collections[data_type]
        
        if import_mode == "overwrite":
            # Clear existing data first
            await collection.delete_many({})
            result = await collection.insert_many(imported_data)
            message = f"Overwritten {len(result.inserted_ids)} {data_type} records"
        else:  # append mode
            result = await collection.insert_many(imported_data)
            message = f"Imported {len(result.inserted_ids)} new {data_type} records"
        
        return {
            "success": True,
            "message": message,
            "imported_count": len(imported_data),
            "import_mode": import_mode,
            "data_type": data_type,
            "imported_at": datetime.utcnow().isoformat(),
            "imported_by": current_user.get("full_name", "Unknown")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error importing data: {str(e)}")

@router.get("/import-options")
async def get_import_options(current_user: dict = Depends(get_current_user)):
    """Get available import options and data types"""
    try:
        # Check permissions
        if current_user.get("role") not in ["admin", "gm", "manager", "accounting"]:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        
        return {
            "data_types": [
                {
                    "id": "clinics",
                    "name": "العيادات",
                    "description": "إدارة بيانات العيادات والأطباء",
                    "icon": "🏥"
                },
                {
                    "id": "users", 
                    "name": "المستخدمين",
                    "description": "إدارة المستخدمين والمندوبين",
                    "icon": "👥"
                },
                {
                    "id": "orders",
                    "name": "الطلبات", 
                    "description": "إدارة طلبات الشراء والمبيعات",
                    "icon": "📋"
                },
                {
                    "id": "debts",
                    "name": "المديونية",
                    "description": "إدارة الديون والذمم المدينة", 
                    "icon": "💰"
                },
                {
                    "id": "payments",
                    "name": "التحصيل",
                    "description": "إدارة المدفوعات والتحصيل",
                    "icon": "💳"
                }
            ],
            "import_modes": [
                {
                    "id": "append",
                    "name": "إضافة البيانات الجديدة",
                    "description": "إضافة البيانات المستوردة مع الاحتفاظ بالبيانات الحالية"
                },
                {
                    "id": "overwrite", 
                    "name": "استبدال جميع البيانات",
                    "description": "حذف البيانات الحالية واستبدالها بالبيانات المستوردة"
                }
            ]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting import options: {str(e)}")